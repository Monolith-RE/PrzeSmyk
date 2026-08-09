import io
import json
import os
import re
import sqlite3
import threading
import time
from datetime import datetime

import geopandas as gpd
import pandas as pd
import requests
import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pyproj import Transformer
from shapely.geometry import Point
from shapely.wkt import loads as wkt_loads

# ==============================================================================
# 1. KONFIGURACJA PROJEKTU "PrzeSmyk"
# ==============================================================================
DB_NAME = "PrzeSmyk_crm.db"
PLIK_SIECI = "sieci_komplet.gpkg"
PLIK_SLUPY = "slupy_komplet.gpkg"
PLIK_WYNIKOWY = "PrzeSmyk_Ranking.xlsx"
STATUS_FILE = "job_status.json"

URL_SIECI = "https://github.com/Monolith-RE/PrzeSmyk/releases/download/v1.0/sieci_komplet.gpkg"
URL_SLUPY = "https://github.com/Monolith-RE/PrzeSmyk/releases/download/v1.0/slupy_komplet.gpkg"

st.set_page_config(
    page_title="PrzeSmyk",
    page_icon="🚙",
    layout="wide",
    initial_sidebar_state="expanded",
)

STREFY_SLUZEBNOSCI = {
    "400kv": 30.0,
    "220kv": 25.0,
    "110kv": 20.0,
    "wysokie": 20.0,
    "najwyższe": 30.0,
    "domyslna": 15.0,
}
WSPOLCZYNNIK_WSPOLKORZYSTANIA = 0.5
CZARNA_LISTA = [
    "osiedle", "os.", "blok", "bloki", "apartament", 
    "apartments", "flats", "wielorodzinny", "spółdzielnia"
]

# ==============================================================================
# 2. STATUS ZADANIA W TLE I BAZA CRM
# ==============================================================================
def set_job_status(status, message="", count=0):
    data = {
        "status": status,
        "message": message,
        "count": count,
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }
    with open(STATUS_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def get_job_status():
    if os.path.exists(STATUS_FILE):
        try:
            with open(STATUS_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {"status": "IDLE", "message": "", "count": 0, "updated_at": ""}

def init_db():
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS historia_dzialek (
                    id_dzialki TEXT PRIMARY KEY, 
                    status TEXT, 
                    nr_kw TEXT, 
                    wlasciciel_dane TEXT, 
                    notatka TEXT, 
                    data_aktualizacji TEXT)""")
    conn.commit()
    conn.close()

def save_crm_record(id_dzialki, status, nr_kw, wlasciciel, notatka):
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    teraz = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    c.execute(
        """INSERT INTO historia_dzialek VALUES (?, ?, ?, ?, ?, ?) 
                 ON CONFLICT(id_dzialki) DO UPDATE SET 
                 status=excluded.status, 
                 nr_kw=excluded.nr_kw, 
                 wlasciciel_dane=excluded.wlasciciel_dane, 
                 notatka=excluded.notatka, 
                 data_aktualizacji=excluded.data_aktualizacji""",
        (id_dzialki, status, nr_kw, wlasciciel, notatka, teraz),
    )
    conn.commit()
    conn.close()

def get_visited_ids():
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute(
        "SELECT id_dzialki FROM historia_dzialek WHERE status IN ('Odwiedzona', 'Finalizacja', 'Odmowa')"
    )
    rows = c.fetchall()
    conn.close()
    return [r[0] for r in rows]

def get_all_crm_records():
    conn = sqlite3.connect(DB_NAME)
    df = pd.read_sql_query(
        "SELECT * FROM historia_dzialek ORDER BY data_aktualizacji DESC", conn
    )
    conn.close()
    return df

def get_val(row, *keys, default=""):
    """Bezpieczny odczyt wartości z wiersza bez względu na nazwę kolumny."""
    for k in keys:
        if k in row and pd.notnull(row[k]) and str(row[k]).strip() != "":
            return row[k]
    return default

@st.cache_data(show_spinner=False)
def pobierz_i_zaladuj_dane_gis():
    if not os.path.exists(PLIK_SIECI):
        r = requests.get(URL_SIECI, stream=True, timeout=60)
        if r.status_code == 200:
            with open(PLIK_SIECI, "wb") as f:
                for chunk in r.iter_content(chunk_size=1024 * 1024):
                    if chunk: f.write(chunk)

    if not os.path.exists(PLIK_SLUPY):
        r = requests.get(URL_SLUPY, stream=True, timeout=60)
        if r.status_code == 200:
            with open(PLIK_SLUPY, "wb") as f:
                for chunk in r.iter_content(chunk_size=1024 * 1024):
                    if chunk: f.write(chunk)

    sieci = gpd.read_file(PLIK_SIECI).to_crs("EPSG:2177")
    slupy = gpd.read_file(PLIK_SLUPY).to_crs("EPSG:2177") if os.path.exists(PLIK_SLUPY) else gpd.GeoDataFrame()
    return sieci, slupy

def generuj_i_zapisz_excel(df_ranking, plik_wyjsciowy=PLIK_WYNIKOWY):
    """Zapisuje kompletny ranking do pliku Excel z 100% klikalnymi hiperłączami openpyxl."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Marszruta PrzeSmyk"

    headers = [
        "Pozycja", "Adres", "Gmina", "ID Działki", "Nr Działki", "Typ Terenu", 
        "Rodzaj Linii", "Pow. Pasa (m²)", "Cena m² (PLN)", "Roszczenie (PLN)", 
        "Liczba Słupów", "Dystans (km)", "Geoportal", "e-Mapa", "OnGeo", "Nawigacja Google",
        "LinkG", "LinkE", "LinkO", "LinkM"
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    link_font = Font(name="Calibri", size=11, color="0000FF", underline="single")

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for idx, row in df_ranking.iterrows():
        r = idx + 2
        id_dz = get_val(row, 'ID', 'ID_Dzialki', 'ID Działki')
        adres_dz = get_val(row, 'Adres')
        gmina_dz = get_val(row, 'Gmina')
        nr_dz = get_val(row, 'Nr', 'Nr_Dzialki', 'Nr Działki')
        typ_dz = get_val(row, 'Typ', 'Typ_Terenu', 'Typ Terenu')
        linia_dz = get_val(row, 'Linia', 'Rodzaj_Linii', 'Rodzaj Linii')
        pow_dz = float(get_val(row, 'Pow', 'Pow_Pasa_m2', default=0.0))
        cena_dz = float(get_val(row, 'Cena', 'Cena_m2_PLN', default=0.0))
        roszczenie_dz = float(get_val(row, 'Roszczenie', 'Roszczenie_PLN', default=0.0))
        slupy_dz = int(get_val(row, 'Slupy', default=0))
        dist_dz = float(get_val(row, 'Dist', default=0.0))

        lg = get_val(row, 'LinkG', 'Link_Geoportal')
        le = get_val(row, 'LinkE', 'Link_Emapa')
        lo = get_val(row, 'LinkO', 'Link_Ongeo')
        lm = get_val(row, 'LinkM', 'Link_Gmaps')

        ws.cell(row=r, column=1, value=idx + 1)
        ws.cell(row=r, column=2, value=str(adres_dz))
        ws.cell(row=r, column=3, value=str(gmina_dz))
        ws.cell(row=r, column=4, value=str(id_dz))
        ws.cell(row=r, column=5, value=str(nr_dz))
        ws.cell(row=r, column=6, value=str(typ_dz))
        ws.cell(row=r, column=7, value=str(linia_dz))
        ws.cell(row=r, column=8, value=pow_dz)
        ws.cell(row=r, column=9, value=cena_dz)
        ws.cell(row=r, column=10, value=roszczenie_dz)
        ws.cell(row=r, column=11, value=slupy_dz)
        ws.cell(row=r, column=12, value=dist_dz)

        # Klikalne linki w Excelu
        links_data = [
            (13, lg, "Otwórz Geoportal"),
            (14, le, "Otwórz e-Mapę"),
            (15, lo, "Otwórz OnGeo"),
            (16, lm, "Nawiguj do działki")
        ]
        for col_i, url, txt in links_data:
            cell = ws.cell(row=r, column=col_i)
            if url and str(url) != 'nan':
                cell.value = f'=HYPERLINK("{url}", "{txt}")'
                cell.hyperlink = str(url)
                cell.font = link_font

        # Surowe adresy URL dla odczytu przez pandas
        ws.cell(row=r, column=17, value=str(lg))
        ws.cell(row=r, column=18, value=str(le))
        ws.cell(row=r, column=19, value=str(lo))
        ws.cell(row=r, column=20, value=str(lm))

    wb.save(plik_wyjsciowy)
    return plik_wyjsciowy

init_db()

# ==============================================================================
# 3. SILNIK GIS
# ==============================================================================
transformer_4326_to_2177 = Transformer.from_crs("EPSG:4326", "EPSG:2177", always_xy=True)
transformer_2177_to_4326 = Transformer.from_crs("EPSG:2177", "EPSG:4326", always_xy=True)

def geokoduj_wpis_startowy(tekst_wpisu):
    if not tekst_wpisu or not tekst_wpisu.strip():
        return 50.0931, 19.9525, "Kraków"
    tekst = tekst_wpisu.strip()
    match = re.search(r"(-?\d+\.\d+)[\s,]+(-?\d+\.\d+)", tekst)
    if match:
        return float(match.group(1)), float(match.group(2)), f"GPS ({float(match.group(1)):.4f}, {float(match.group(2)):.4f})"
    try:
        url = f"https://nominatim.openstreetmap.org/search?q={requests.utils.quote(tekst)}&format=json&limit=1"
        r = requests.get(url, headers={"User-Agent": "PrzeSmykApp/14.0"}, timeout=5)
        if r.status_code == 200 and len(r.json()) > 0:
            res = r.json()[0]
            return float(res["lat"]), float(res["lon"]), res.get("display_name", tekst).split(",")[0]
    except Exception:
        pass
    return 50.0931, 19.9525, "Kraków (Domyślnie)"

def pobierz_adres_i_filtr_zabudowy(lat, lon, nr_dzialki_ewidencja):
    try:
        url = f"https://nominatim.openstreetmap.org/reverse?format=json&lat={lat}&lon={lon}&zoom=18&addressdetails=1&extratags=1"
        r = requests.get(url, headers={"User-Agent": "PrzeSmykApp/14.0"}, timeout=5)
        if r.status_code == 200:
            data = r.json()
            raw_text = str(data).lower()

            if any(s in raw_text for s in CZARNA_LISTA):
                return "", False, "Osiedle / Bloki"
            if data.get("extratags", {}).get("building") in ["apartments", "residential", "dormitory", "terrace"]:
                return "", False, "Wielorodzinny"

            addr = data.get("address", {})
            miasto = addr.get("city") or addr.get("town") or addr.get("village") or "Kraków"
            ulica = addr.get("road") or ""
            numer = addr.get("house_number") or ""

            adres_czysty = (
                f"{miasto}, ul. {ulica} {numer}".strip()
                if ulica and numer
                else f"{miasto}, dz. nr {nr_dzialki_ewidencja}"
            )

            if any(b in raw_text for b in ["commercial", "industrial", "warehouse", "company"]):
                return adres_czysty, True, "🏭 Firma / Przemysł"
            return adres_czysty, True, "🏠 Dom Jednorodzinny / Posesja"
    except Exception:
        pass
    return f"Kraków, dz. nr {nr_dzialki_ewidencja}", True, "🏠 Posesja"

def uldk_pobierz_dzialke_z_geometria(x_2177, y_2177):
    url = f"https://uldk.gugik.gov.pl/request.php?request=GetParcelByXY&xy={x_2177},{y_2177},2177&result=id,commune,parcel,geom_wkt"
    try:
        resp = requests.get(url, timeout=5)
        if resp.status_code == 200 and not resp.text.startswith("-1"):
            lines = resp.text.strip().split("\n")
            if len(lines) >= 2:
                p = lines[1].split("|")
                id_d = p[0]
                gmina = p[1]
                nr_d = p[2]
                wkt_str = p[3] if len(p) > 3 else None
                geom = wkt_loads(wkt_str) if wkt_str else None
                return {
                    "id_dzialki": id_d,
                    "gmina": gmina,
                    "nr_dzialki": nr_d,
                    "geom": geom,
                    "x": x_2177,
                    "y": y_2177,
                }
    except Exception:
        pass
    return None

def szacuj_cene_m2_avm(odleglosc_dom_km):
    return max(180.0, 600.0 - (odleglosc_dom_km * 8.0))

# ==============================================================================
# 4. PRACA W TLE NA SERWERZE
# ==============================================================================
def worker_przelicz_trase(current_lat, current_lon):
    try:
        set_job_status("RUNNING", "Skanowanie terenu i granic działek w głąb powiatów...")
        
        sieci, slupy = pobierz_i_zaladuj_dane_gis()
        
        dom_x, dom_y = transformer_4326_to_2177.transform(current_lon, current_lat)
        punkt_dom = Point(dom_x, dom_y)
        odwiedzone_ids = get_visited_ids()

        sieci_work = sieci.copy()
        sieci_work["dist"] = sieci_work.geometry.distance(punkt_dom) / 1000.0
        sieci_work = sieci_work.sort_values(by="dist", ascending=True)

        TARGET_COUNT = 100
        wykryte_dzialki = {}
        PROMIENIE = [5000.0, 15000.0, 30000.0, 50000.0]

        for promien in PROMIENIE:
            if len(wykryte_dzialki) >= TARGET_COUNT:
                break

            bufor_obszaru = punkt_dom.buffer(promien)
            sieci_w_promieniu = sieci_work[sieci_work.geometry.intersects(bufor_obszaru)].copy()

            for idx, linia in sieci_w_promieniu.iterrows():
                if len(wykryte_dzialki) >= TARGET_COUNT:
                    break

                opis_nap = str(linia.get("napiecie", linia.get("rodzaj", "110 kV"))).upper()
                szerokosc_strefy = 15.0
                for k, v in STREFY_SLUZEBNOSCI.items():
                    if k in opis_nap.lower():
                        szerokosc_strefy = v
                        break

                dlugosc = linia.geometry.length
                for d in range(0, int(dlugosc), 120):
                    if len(wykryte_dzialki) >= TARGET_COUNT:
                        break

                    pt = linia.geometry.interpolate(d)
                    dzialka = uldk_pobierz_dzialke_z_geometria(pt.x, pt.y)
                    if dzialka:
                        id_d = dzialka["id_dzialki"]
                        if id_d in odwiedzone_ids or id_d in wykryte_dzialki:
                            continue

                        poly = dzialka.get("geom")
                        if poly and not poly.is_empty:
                            rdzen_dzialki = poly.buffer(-2.5)
                            if rdzen_dzialki.is_empty or not rdzen_dzialki.intersects(linia.geometry):
                                continue

                        lon_wgs, lat_wgs = transformer_2177_to_4326.transform(pt.x, pt.y)
                        adres_czysty, ok, typ_terenu = pobierz_adres_i_filtr_zabudowy(lat_wgs, lon_wgs, dzialka["nr_dzialki"])

                        if not ok:
                            continue

                        dzialka.update({
                            "szer_pasa": szerokosc_strefy,
                            "rodzaj": opis_nap,
                            "adres": adres_czysty,
                            "typ": typ_terenu,
                            "lat": lat_wgs,
                            "lon": lon_wgs,
                            "dist": linia["dist"],
                        })
                        wykryte_dzialki[id_d] = dzialka
                        set_job_status("RUNNING", f"Znaleziono {len(wykryte_dzialki)}/100 działek...", len(wykryte_dzialki))
                        time.sleep(0.01)

        lista_rankingowa = []
        for id_d, d in wykryte_dzialki.items():
            ilosc_slupow = (
                len(slupy[slupy.geometry.intersects(Point(d["x"], d["y"]).buffer(30.0))])
                if not slupy.empty else 0
            )
            cena = szacuj_cene_m2_avm(d["dist"])
            pow_pasa = 85.0 * d["szer_pasa"]
            roszczenie = pow_pasa * cena * WSPOLCZYNNIK_WSPOLKORZYSTANIA

            link_geo = f"https://mapy.geoportal.gov.pl/imap/Imgp_2.html?identifyParcel={id_d}"
            link_ema = f"https://polska.e-mapa.net/?dzialka={id_d}"
            link_ong = f"https://ongeo.pl/mapa?x={d['lon']:.6f}&y={d['lat']:.6f}&zoom=19"
            link_gmaps = f"https://www.google.com/maps?q={d['lat']},{d['lon']}"

            rodzaj_mediow = "🔥 Gazociąg" if "GAZ" in d["rodzaj"] else "⚡ Linia Elektroenergetyczna"

            lista_rankingowa.append({
                "ID": id_d,
                "Adres": d["adres"],
                "Gmina": d["gmina"],
                "Nr": d["nr_dzialki"],
                "Typ": d["typ"],
                "Linia": f"{rodzaj_mediow} ({d['rodzaj']})",
                "Pow": pow_pasa,
                "Cena": cena,
                "Roszczenie": round(roszczenie, 2),
                "Slupy": ilosc_slupow,
                "Dist": round(d["dist"], 2),
                "LinkG": link_geo,
                "LinkE": link_ema,
                "LinkO": link_ong,
                "LinkM": link_gmaps,
            })

        df = pd.DataFrame(lista_rankingowa)
        if not df.empty:
            df = df.sort_values(by=["Roszczenie", "Dist"], ascending=[False, True]).reset_index(drop=True)
            generuj_i_zapisz_excel(df, PLIK_WYNIKOWY)
            set_job_status("COMPLETED", f"Pomyślnie wygenerowano marszrutę dla {len(df)} działek!", len(df))
        else:
            set_job_status("ERROR", "Brak działek spełniających kryteria w wybranym obszarze.")

    except Exception as e:
        set_job_status("ERROR", f"Błąd podczas pracy w tle: {str(e)}")

# ==============================================================================
# 5. INTERFEJS STREAMLIT
# ==============================================================================
st.sidebar.title("🚙⚡🔥 PrzeSmyk v3.1")
st.sidebar.caption("Ranking służebności i planowanie trasy")
st.sidebar.markdown("---")

st.sidebar.subheader("📍 Start Marszruty")
wpis_lokalizacji = st.sidebar.text_input(
    "Wpisz adres startowy:",
    value="Kraków, ul. Nad Sudołem 32",
)
current_lat, current_lon, opis_lokalizacji = geokoduj_wpis_startowy(wpis_lokalizacji)
st.sidebar.caption(f"🎯 Zlokalizowano: **{opis_lokalizacji}**")

przelicz_button = st.sidebar.button("🚙 PRZELICZ TRASĘ W TLE", type="primary")

tab1, tab2, tab3 = st.tabs(
    ["🗺️ Trasa & Operat", "📝 CRM Terenowy", "🗂️ Baza Działek"]
)

if przelicz_button:
    if os.path.exists(PLIK_WYNIKOWY):
        try: os.remove(PLIK_WYNIKOWY)
        except Exception: pass
    if 'rank' in st.session_state:
        del st.session_state['rank']
        
    set_job_status("RUNNING", "Inicjalizacja nowej analizy w tle...")
    t = threading.Thread(target=worker_przelicz_trase, args=(current_lat, current_lon))
    t.daemon = True
    t.start()
    st.sidebar.success("🚀 Uruchomiono nową analizę! Możesz zamknąć przeglądarkę i wyłączyć iPada.")

with tab1:
    job = get_job_status()
    
    if job.get("status") == "RUNNING":
        st.info(f"⏳ **Trwa analiza w tle na serwerze...**\n\n*Status:* {job.get('message')}\n\n*Ostatnia aktualizacja:* {job.get('updated_at')}\n\n💡 **Możesz bezpiecznie wyłączyć iPada i zamknąć stronę.** Serwer dokończy obliczenia, a plik Excel będzie czekał po ponownym otwarciu strony.")
        if st.button("🔄 Odśwież status"):
            st.rerun()

    elif job.get("status") == "COMPLETED" and os.path.exists(PLIK_WYNIKOWY):
        st.success(f"🎉 **Gotowe!** Nowy raport został pomyślnie wygenerowany ({job.get('updated_at')}).")
        
        with open(PLIK_WYNIKOWY, "rb") as f:
            st.download_button(
                label="📊 Pobierz pełny raport (.XLSX) dla Google Sheets / MS Excel",
                data=f.read(),
                file_name=PLIK_WYNIKOWY,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary"
            )
        st.markdown("---")

        try:
            df_rank = pd.read_excel(PLIK_WYNIKOWY)
            st.session_state["rank"] = df_rank
            st.subheader(f"📍 Wygenerowana Marszruta Terenowa ({len(df_rank)} działek)")

            for idx, row in df_rank.iterrows():
                id_dz = get_val(row, 'ID', 'ID_Dzialki', 'ID Działki')
                adres_dz = get_val(row, 'Adres')
                gmina_dz = get_val(row, 'Gmina')
                typ_dz = get_val(row, 'Typ', 'Typ_Terenu', 'Typ Terenu')
                roszczenie_dz = float(get_val(row, 'Roszczenie', 'Roszczenie_PLN', default=0.0))
                slupy_dz = int(get_val(row, 'Slupy', default=0))
                pow_dz = float(get_val(row, 'Pow', 'Pow_Pasa_m2', default=0.0))
                linia_dz = get_val(row, 'Linia', 'Rodzaj_Linii')

                link_g = get_val(row, 'LinkG', 'Link_Geoportal', 'Geoportal Link', default=f"https://mapy.geoportal.gov.pl/imap/Imgp_2.html?identifyParcel={id_dz}")
                link_e = get_val(row, 'LinkE', 'Link_Emapa', 'e-Mapa Link', default=f"https://polska.e-mapa.net/?dzialka={id_dz}")
                link_o = get_val(row, 'LinkO', 'Link_Ongeo', 'OnGeo Link', default="https://ongeo.pl")
                link_m = get_val(row, 'LinkM', 'Link_Gmaps', 'Nawigacja Google', default="https://maps.google.com")

                st.markdown(f"### {idx+1}. {adres_dz}")
                c1, c2, c3 = st.columns([3, 3, 3])
                c1.markdown(
                    f"📍 **Gmina:** {gmina_dz}\n🆔 **ID Działki:** `{id_dz}`\n{typ_dz}"
                )
                c2.markdown(
                    f"💰 **Roszczenie:** `{roszczenie_dz:,.2f} PLN`\n🗼 **Wieża / Słup:** `{slupy_dz} szt.`\n📐 **Pas:** `{pow_dz} m²`"
                )
                
                c3.markdown(
                    f"🌐 **Weryfikacja Geodezyjna:**\n"
                    f"• [Otwórz Działkę - Geoportal]({link_g})\n"
                    f"• [Otwórz Działkę - e-Mapa]({link_e})\n"
                    f"• [Otwórz Mapę - OnGeo.pl]({link_o})"
                )

                b1, b2, b3 = st.columns([3, 3, 3])
                with b1.popover("📄 Szybki podgląd"):
                    st.write(f"ID: {id_dz}\nAdres: {adres_dz}\nKwalifikacja: {typ_dz}")
                with b2.popover("📊 Kalkulator"):
                    st.write(f"**Infrastruktura:** {linia_dz}\n### Wartość: {roszczenie_dz:,.2f} PLN")
                b3.link_button("🚗 Nawiguj (Google Maps)", link_m, type="primary")
                st.divider()
        except Exception as e:
            st.error(f"Błąd odczytu pliku wyników: {e}")

    elif job.get("status") == "ERROR":
        st.error(f"❌ Wystąpił błąd podczas analizy: {job.get('message')}")
    else:
        st.info("Wpisz adres startowy w panelu po lewej stronie i kliknij **'🚙 PRZELICZ TRASĘ W TLE'**.")

with tab2:
    with st.form("crm"):
        tid = st.text_input("ID Działki")
        stat = st.selectbox(
            "Status Wizyty", ["Odwiedzona", "Umówione spotkanie", "Odmowa"]
        )
        kw = st.text_input("Numer KW")
        wlas = st.text_input("Dane Kontaktowe Właściciela")
        notat = st.text_area("Notatka z rozmowy")
        if st.form_submit_button("💾 Zapisz w CRM"):
            save_crm_record(tid, stat, kw, wlas, notat)
            st.success("Zapisano rekord w bazie PrzeSmyk!")

with tab3:
    st.dataframe(get_all_crm_records(), width="stretch")
